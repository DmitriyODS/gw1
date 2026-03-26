import io
from collections import defaultdict
from datetime import datetime, timedelta
from flask import Blueprint, render_template, send_file, request, jsonify, current_app
from flask_login import login_required, current_user
from sqlalchemy import func
from extensions import db
from models import Task, TimeLog, User, Department, TaskStatus, Urgency

analytics_bp = Blueprint('analytics', __name__)

TYPE_LABELS = {
    'pub_images':           'Картинки для публикаций',
    'banner':               'Разработка баннера',
    'poster':               'Разработка афиши',
    'presentation':         'Разработка презентации',
    'presentation_update':  'Доработка презентации',
    'text_writing':         'Написание текста',
    'handout':              'Разработка раздатки',
    'placement':            'Размещение материалов',
    'internal':             'Внутренняя работа',
    'external':             'Внешняя работа',
    'water_plants':         'Полить цветы',
    'exports':              'Подготовка выгрузок',
    'surveys':              'Создание опросов',
    'photo_edit':           'Обработка фото',
    'video_edit':           'Монтаж ролика',
    'video_shoot':          'Съёмка ролика',
    'photo_shoot':          'Фотосъёмка',
    'meeting':              'Планёрка',
    'mail_work':            'Работа с почтой',
    'cloud_work':           'Работа с облаками',
    'stand_design':         'Разработка стендов',
    'pub_design':           'Разработка дизайна для публикаций',
    'branded':              'Разработка брендированной продукции',
    'small_design':         'Мелкий дизайн',
    # backward compat
    'publication':          'Публикация (устар.)',
    'picture':              'Разработка картинки (устар.)',
    'merch':                'Сувенирная продукция (устар.)',
    'other':                'Другое',
    None:                   'Не указан',
}


def get_range(period):
    from flask import current_app
    tz_hours = current_app.config.get('TZ_OFFSET_HOURS', 3)
    now_utc = datetime.utcnow()
    now_local = now_utc + timedelta(hours=tz_hours)
    if period == 'day':
        start_local = now_local.replace(hour=0, minute=0, second=0, microsecond=0)
    elif period == 'week':
        start_local = (now_local - timedelta(days=now_local.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
    elif period == 'month':
        start_local = now_local.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    elif period == 'year':
        start_local = now_local.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    else:
        start_local = now_local - timedelta(days=7)
    start = start_local - timedelta(hours=tz_hours)
    return start, now_utc


def build_stats(period):
    start, end = get_range(period)

    # 1 запрос: количество задач по статусам (вместо 5 отдельных COUNT)
    status_rows = db.session.query(Task.status, func.count(Task.id)).filter(
        Task.created_at >= start
    ).group_by(Task.status).all()
    by_status = {s: 0 for s in TaskStatus.LABELS}
    for s, cnt in status_rows:
        if s in by_status:
            by_status[s] = cnt

    # 1 запрос: тип × статус (вместо N_типов × 5 запросов)
    type_status_rows = db.session.query(
        Task.task_type, Task.status, func.count(Task.id)
    ).filter(Task.created_at >= start).group_by(Task.task_type, Task.status).all()

    type_totals = defaultdict(int)
    _type_by_status_raw = defaultdict(lambda: defaultdict(int))
    for ttype, s, cnt in type_status_rows:
        type_totals[ttype] += cnt
        if s in TaskStatus.LABELS:
            _type_by_status_raw[ttype][s] = cnt

    raw_type_stats = sorted(type_totals.items(), key=lambda x: x[1], reverse=True)
    type_stats = [(TYPE_LABELS.get(t, t or 'Не указан'), cnt) for t, cnt in raw_type_stats]
    type_by_status = {}
    for ttype, _ in raw_type_stats:
        label = TYPE_LABELS.get(ttype, ttype or 'Другое')
        type_by_status[label] = {s: _type_by_status_raw[ttype].get(s, 0) for s in TaskStatus.LABELS}

    # 1 запрос: отдел × статус (вместо 8 × 5 = 40 запросов)
    dept_status_rows = db.session.query(
        Department.name, Task.status, func.count(Task.id)
    ).join(Task, Task.department_id == Department.id).filter(
        Task.created_at >= start
    ).group_by(Department.name, Task.status).all()

    dept_totals = defaultdict(int)
    _dept_by_status_raw = defaultdict(lambda: defaultdict(int))
    for dname, s, cnt in dept_status_rows:
        dept_totals[dname] += cnt
        if s in TaskStatus.LABELS:
            _dept_by_status_raw[dname][s] = cnt

    dept_stats = sorted(dept_totals.items(), key=lambda x: x[1], reverse=True)
    dept_by_status = {}
    for dname, _ in dept_stats[:8]:
        dept_by_status[dname] = {s: _dept_by_status_raw[dname].get(s, 0) for s in TaskStatus.LABELS}

    # top_dept_incoming выводится из уже вычисленного dept_stats — запрос не нужен
    top_dept_incoming = dept_stats[:10]

    top_customers = db.session.query(
        Task.customer_name, func.count(Task.id).label('cnt')
    ).filter(
        Task.created_at >= start, Task.customer_name != None, Task.customer_name != ''
    ).group_by(Task.customer_name).order_by(func.count(Task.id).desc()).limit(10).all()

    top_tasks = db.session.query(User.full_name, func.count(Task.id).label('cnt')).join(
        Task, Task.created_by_id == User.id
    ).filter(Task.status == TaskStatus.DONE, Task.created_at >= start).group_by(
        User.full_name
    ).order_by(func.count(Task.id).desc()).limit(10).all()

    top_time = db.session.query(
        User.full_name,
        func.coalesce(func.sum(
            func.extract('epoch', TimeLog.ended_at - TimeLog.started_at)
        ), 0).label('secs')
    ).join(TimeLog, TimeLog.user_id == User.id).filter(
        TimeLog.started_at >= start, TimeLog.ended_at != None
    ).group_by(User.full_name).order_by(func.sum(
        func.extract('epoch', TimeLog.ended_at - TimeLog.started_at)
    ).desc()).limit(10).all()

    return {
        'by_status': by_status,
        'dept_stats': dept_stats,
        'type_stats': type_stats,
        'type_by_status': type_by_status,
        'dept_by_status': dept_by_status,
        'top_customers': top_customers,
        'top_dept_incoming': top_dept_incoming,
        'top_tasks': top_tasks,
        'top_time': top_time,
    }


def _period_bounds(mode, offset, tz_hours):
    """Return (start_utc, end_utc, label) for given mode/offset."""
    import calendar as cal_mod
    now_utc = datetime.utcnow()
    now_local = now_utc + timedelta(hours=tz_hours)

    if mode == 'day':
        day = (now_local + timedelta(days=offset)).date()
        s = datetime(day.year, day.month, day.day, 0, 0, 0)
        e = datetime(day.year, day.month, day.day, 23, 59, 59)
        label = day.strftime('%d.%m.%Y')
    elif mode == 'week':
        monday = (now_local - timedelta(days=now_local.weekday())).date() + timedelta(weeks=offset)
        sunday = monday + timedelta(days=6)
        s = datetime(monday.year, monday.month, monday.day, 0, 0, 0)
        e = datetime(sunday.year, sunday.month, sunday.day, 23, 59, 59)
        label = f'{monday.strftime("%d.%m")} – {sunday.strftime("%d.%m.%Y")}'
    else:  # month
        year = now_local.year
        month = now_local.month + offset
        while month <= 0:
            month += 12; year -= 1
        while month > 12:
            month -= 12; year += 1
        last_day = cal_mod.monthrange(year, month)[1]
        s = datetime(year, month, 1, 0, 0, 0)
        e = datetime(year, month, last_day, 23, 59, 59)
        label = s.strftime('%B %Y')

    return s - timedelta(hours=tz_hours), e - timedelta(hours=tz_hours), label


@analytics_bp.route('/analytics/time')
@login_required
def time_report():
    can_see_all = current_user.can_admin
    mode   = request.args.get('mode', 'week')
    offset = request.args.get('offset', 0, type=int)
    tz_hours = current_app.config.get('TZ_OFFSET_HOURS', 3)
    start, end, period_label = _period_bounds(mode, offset, tz_hours)

    all_users = (User.query
                 .filter_by(is_active=True)
                 .filter(User.role != 'tv')
                 .order_by(User.full_name)
                 .all()) if can_see_all else [current_user]

    # 1 запрос: суммарное время по пользователям (вместо N запросов)
    secs_q = db.session.query(
        TimeLog.user_id,
        func.coalesce(func.sum(
            func.extract('epoch', TimeLog.ended_at - TimeLog.started_at)
        ), 0).label('secs')
    ).filter(
        TimeLog.started_at >= start,
        TimeLog.started_at <= end,
        TimeLog.ended_at.isnot(None),
    ).group_by(TimeLog.user_id).all()
    secs_map = {r.user_id: int(r.secs) for r in secs_q}

    # 1 запрос: закрытые задачи по пользователям (вместо N запросов)
    done_q = db.session.query(
        TimeLog.user_id,
        func.count(func.distinct(Task.id)).label('cnt')
    ).join(Task, TimeLog.task_id == Task.id).filter(
        TimeLog.ended_at.isnot(None),
        Task.status == TaskStatus.DONE,
        Task.completed_at >= start,
        Task.completed_at <= end,
        Task.completed_at.isnot(None),
    ).group_by(TimeLog.user_id).all()
    done_map = {r.user_id: r.cnt for r in done_q}

    rows = [
        {'id': u.id, 'name': u.full_name,
         'done': done_map.get(u.id, 0),
         'secs': secs_map.get(u.id, 0)}
        for u in all_users
    ]

    all_scores = compute_staff_scores(start, end, limit=None)
    score_map = {s['id']: {'score': s['score'], 'rank': i + 1} for i, s in enumerate(all_scores)}
    total_ranked = len(all_scores)
    for row in rows:
        info = score_map.get(row['id'])
        row['score'] = info['score'] if info else None
        row['rank'] = info['rank'] if info else None

    rows.sort(key=lambda r: (
        0 if r['score'] is not None else 1,  # с рейтингом — выше
        -(r['score'] or 0),
        -r['done'],
        -r['secs'],
    ))

    today_iso = (datetime.utcnow() + timedelta(hours=tz_hours)).strftime('%Y-%m-%d')
    return render_template('analytics/time.html',
        mode=mode, offset=offset, period_label=period_label,
        can_see_all=can_see_all, rows=rows, total_ranked=total_ranked, today_iso=today_iso)


@analytics_bp.route('/analytics/time/user-detail')
@login_required
def time_user_detail():
    """Returns JSON detail for one user in a given period."""
    can_see_all = current_user.can_admin
    uid    = request.args.get('user_id', type=int)
    mode   = request.args.get('mode', 'week')
    offset = request.args.get('offset', 0, type=int)
    tz_hours = current_app.config.get('TZ_OFFSET_HOURS', 3)

    if not uid:
        uid = current_user.id
    if not can_see_all and uid != current_user.id:
        return jsonify({'error': 'Нет прав'}), 403

    start, end, period_label = _period_bounds(mode, offset, tz_hours)

    worked_task_ids = db.session.query(TimeLog.task_id).filter(
        TimeLog.user_id == uid,
        TimeLog.ended_at.isnot(None),
    ).distinct().subquery()
    done_tasks = Task.query.filter(
        Task.id.in_(worked_task_ids),
        Task.status == TaskStatus.DONE,
        Task.completed_at >= start,
        Task.completed_at <= end,
        Task.completed_at.isnot(None),
    ).order_by(Task.completed_at.desc()).all()

    logs = TimeLog.query.filter(
        TimeLog.user_id == uid,
        TimeLog.started_at >= start,
        TimeLog.started_at <= end,
        TimeLog.ended_at.isnot(None),
    ).all()
    time_by_task = defaultdict(float)
    for log in logs:
        time_by_task[log.task_id] += (log.ended_at - log.started_at).total_seconds()

    total_secs = int(sum(time_by_task.values()))

    by_type = defaultdict(lambda: {'cnt': 0, 'secs': 0})
    tasks_out = []
    for t in done_tasks:
        secs = int(time_by_task.get(t.id, 0))
        type_label = TYPE_LABELS.get(t.task_type, t.task_type or 'Не указан')
        by_type[type_label]['cnt'] += 1
        by_type[type_label]['secs'] += secs
        tasks_out.append({
            'id': t.id,
            'title': t.title,
            'type': type_label,
            'secs': secs,
            'completed_at': (t.completed_at + timedelta(hours=tz_hours)).strftime('%d.%m %H:%M'),
        })

    type_list = sorted(
        [{'label': k, 'cnt': v['cnt'], 'secs': v['secs']} for k, v in by_type.items()],
        key=lambda x: x['cnt'], reverse=True
    )

    user = User.query.get(uid)
    return jsonify({
        'user_name': user.full_name if user else '',
        'period_label': period_label,
        'total_tasks': len(done_tasks),
        'total_secs': total_secs,
        'tasks': tasks_out,
        'by_type': type_list,
    })


def build_burnup(period):
    from flask import current_app
    tz_hours = current_app.config.get('TZ_OFFSET_HOURS', 3)
    start, _ = get_range(period)
    tz_delta = timedelta(hours=tz_hours)
    now_utc = datetime.utcnow()
    now_local_date = (now_utc + tz_delta).date()

    # 1 запрос: созданные задачи по дням (GROUP BY в SQL вместо Python-цикла по всем задачам)
    created_date_expr = func.date(Task.created_at + tz_delta)
    created_rows = db.session.query(
        created_date_expr.label('d'), func.count(Task.id)
    ).filter(Task.created_at >= start).group_by(created_date_expr).all()

    # 1 запрос: закрытые задачи по дням
    done_date_expr = func.date(Task.completed_at + tz_delta)
    done_rows = db.session.query(
        done_date_expr.label('d'), func.count(Task.id)
    ).filter(
        Task.status == TaskStatus.DONE,
        Task.completed_at >= start,
        Task.completed_at.isnot(None)
    ).group_by(done_date_expr).all()

    days_data = defaultdict(lambda: {'created': 0, 'done': 0})
    for d, cnt in created_rows:
        days_data[d]['created'] = cnt
    for d, cnt in done_rows:
        days_data[d]['done'] = cnt

    dates, c_series, d_series = [], [], []
    c_c = c_d = 0
    cur = (start + tz_delta).date()
    while cur <= now_local_date:
        c_c += days_data[cur]['created']
        c_d += days_data[cur]['done']
        dates.append(cur.strftime('%d.%m'))
        c_series.append(c_c)
        d_series.append(c_d)
        cur += timedelta(days=1)
    return {'dates': dates, 'created': c_series, 'done': d_series}


@analytics_bp.route('/analytics')
@login_required
def dashboard():
    period = request.args.get('period', 'week')
    stats = build_stats(period)
    burnup = build_burnup(period)
    return render_template('analytics/dashboard.html', period=period,
                           TaskStatus=TaskStatus, Urgency=Urgency,
                           burnup=burnup, **stats)


@analytics_bp.route('/analytics/tv')
def tv_mode():
    return render_template('analytics/tv.html')


def compute_staff_scores(start, end=None, limit=5):
    """Compute employee ranking using formula R_i = 0.5*(N_i/N_max) + 0.5*(S_i/S_max), scaled to 0-100.
    N_i — кол-во закрытых задач в периоде, S_i — суммарное время работы по этим задачам."""
    now = datetime.utcnow()
    if end is None:
        end = now

    user_tasks_q = db.session.query(
        TimeLog.user_id,
        User.full_name,
        func.count(func.distinct(Task.id)).label('cnt')
    ).join(Task, TimeLog.task_id == Task.id).join(User, TimeLog.user_id == User.id).filter(
        Task.status == TaskStatus.DONE,
        Task.completed_at >= start,
        Task.completed_at <= end,
        Task.completed_at.isnot(None),
        TimeLog.ended_at.isnot(None),
    ).group_by(TimeLog.user_id, User.full_name).all()

    # Считаем суммарное время только по закрытым задачам того же периода,
    # чтобы периоды задач и времени совпадали.
    user_time_q = db.session.query(
        TimeLog.user_id,
        func.coalesce(func.sum(
            func.extract('epoch', TimeLog.ended_at - TimeLog.started_at)
        ), 0).label('secs')
    ).join(Task, TimeLog.task_id == Task.id).filter(
        Task.status == TaskStatus.DONE,
        Task.completed_at >= start,
        Task.completed_at <= end,
        Task.completed_at.isnot(None),
        TimeLog.ended_at.isnot(None),
    ).group_by(TimeLog.user_id).all()

    time_map = {u.user_id: float(u.secs) for u in user_time_q}

    staff = []
    for u in user_tasks_q:
        n_i = u.cnt
        total_secs = time_map.get(u.user_id, 0)
        # t = суммарное время (больше = лучше), не среднее
        staff.append({'id': u.user_id, 'name': u.full_name, 'n': n_i, 't': total_secs, 'secs': int(total_secs)})

    if not staff:
        return []

    n_max = max(s['n'] for s in staff) or 1
    t_max = max(s['t'] for s in staff) or 1

    for s in staff:
        s['score'] = round((0.5 * s['n'] / n_max + 0.5 * s['t'] / t_max) * 100)

    ranked = sorted(staff, key=lambda x: x['score'], reverse=True)
    return ranked[:limit] if limit is not None else ranked


@analytics_bp.route('/analytics/tv/data')
def tv_data():
    now = datetime.utcnow()
    tz_hours = current_app.config.get('TZ_OFFSET_HOURS', 3)
    now_local = now + timedelta(hours=tz_hours)

    def period_data(start, end=None):
        if end is None:
            end = now
        # 1 запрос вместо 4 отдельных COUNT
        status_rows = db.session.query(Task.status, func.count(Task.id)).filter(
            Task.created_at >= start,
            Task.created_at <= end,
            Task.is_archived == False,
        ).group_by(Task.status).all()
        status = {s: 0 for s in ['new', 'in_progress', 'paused', 'done']}
        for s, cnt in status_rows:
            if s in status:
                status[s] = cnt

        raw_types = db.session.query(Task.task_type, func.count(Task.id)).filter(
            Task.created_at >= start, Task.created_at <= end
        ).group_by(Task.task_type).order_by(func.count(Task.id).desc()).all()
        types = [{'label': TYPE_LABELS.get(t, t or 'Другое'), 'cnt': c} for t, c in raw_types if c > 0]

        total_secs = int(db.session.query(
            func.coalesce(func.sum(
                func.extract('epoch', TimeLog.ended_at - TimeLog.started_at)
            ), 0)
        ).filter(TimeLog.started_at >= start, TimeLog.started_at <= end, TimeLog.ended_at.isnot(None)).scalar() or 0)

        dept_q = db.session.query(Department.name, func.count(Task.id)).join(
            Task, Task.department_id == Department.id
        ).filter(Task.created_at >= start, Task.created_at <= end
        ).group_by(Department.name).order_by(func.count(Task.id).desc()).limit(8).all()
        depts = [{'name': n, 'cnt': c} for n, c in dept_q]

        staff = compute_staff_scores(start, end)

        return {
            'status': status,
            'types': types,
            'total_secs': total_secs,
            'depts': depts,
            'staff': [{'name': s['name'], 'score': s['score'], 'tasks': s['n'], 'secs': s['secs']} for s in staff],
        }

    today_start = datetime(now_local.year, now_local.month, now_local.day, 0, 0, 0) - timedelta(hours=tz_hours)
    today_data = period_data(today_start)

    monday_local = (now_local - timedelta(days=now_local.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
    week_start = monday_local - timedelta(hours=tz_hours)
    week_data = period_data(week_start)

    # 1 запрос вместо 3 отдельных COUNT для alltime-статусов
    alltime_rows = db.session.query(Task.status, func.count(Task.id)).group_by(Task.status).all()
    alltime_status_map = defaultdict(int)
    for s, cnt in alltime_rows:
        alltime_status_map[s] = cnt
    alltime_new  = alltime_status_map.get('new', 0)
    alltime_ip   = alltime_status_map.get('in_progress', 0) + alltime_status_map.get('paused', 0)
    alltime_done = alltime_status_map.get('done', 0)

    alltime_types_q = db.session.query(Task.task_type, func.count(Task.id)).group_by(Task.task_type).order_by(func.count(Task.id).desc()).all()
    alltime_types = [{'label': TYPE_LABELS.get(t, t or 'Другое'), 'cnt': c} for t, c in alltime_types_q if c > 0]
    alltime_depts_q = db.session.query(Department.name, func.count(Task.id)).join(
        Task, Task.department_id == Department.id
    ).group_by(Department.name).order_by(func.count(Task.id).desc()).limit(8).all()
    alltime_depts = [{'name': n, 'cnt': c} for n, c in alltime_depts_q]
    year_start = datetime(now_local.year, 1, 1, 0, 0, 0) - timedelta(hours=tz_hours)
    alltime_staff = compute_staff_scores(year_start)

    return jsonify({
        'today': today_data,
        'week': week_data,
        'alltime': {
            'new': alltime_new,
            'in_progress': alltime_ip,
            'done': alltime_done,
            'types': alltime_types,
            'depts': alltime_depts,
            'staff': [{'name': s['name'], 'score': s['score'], 'tasks': s['n']} for s in alltime_staff],
        },
    })


@analytics_bp.route('/analytics/export/excel')
@login_required
def export_excel():
    import openpyxl
    from openpyxl.styles import Font
    from sqlalchemy.orm import joinedload

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Задачи'
    headers = ['ID', 'Заголовок', 'Тип', 'Статус', 'Срочность', 'Отдел', 'Заказчик',
               'Телефон', 'Дедлайн', 'Создана', 'Время (мин)']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Суммарное время на задачу через SQL (вместо Python-цикла по time_logs для каждой задачи)
    secs_sq = db.session.query(
        TimeLog.task_id,
        func.coalesce(func.sum(
            func.extract('epoch', TimeLog.ended_at - TimeLog.started_at)
        ), 0).label('secs')
    ).filter(TimeLog.ended_at.isnot(None)).group_by(TimeLog.task_id).subquery()

    rows = (db.session.query(Task, func.coalesce(secs_sq.c.secs, 0))
            .outerjoin(secs_sq, Task.id == secs_sq.c.task_id)
            .options(joinedload(Task.department))
            .order_by(Task.created_at.desc())
            .all())

    for t, total_secs in rows:
        ws.append([
            t.id, t.title,
            TYPE_LABELS.get(t.task_type, t.task_type or ''),
            TaskStatus.LABELS.get(t.status, t.status),
            Urgency.LABELS.get(t.urgency, t.urgency),
            t.department.name if t.department else '',
            t.customer_name or '', t.customer_phone or '',
            t.deadline.strftime('%d.%m.%Y %H:%M') if t.deadline else '',
            t.created_at.strftime('%d.%m.%Y %H:%M'),
            round(total_secs / 60),
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='tasks_export.xlsx')


@analytics_bp.route('/analytics/export/user-stats-excel')
@login_required
def export_user_stats_excel():
    import openpyxl
    from openpyxl.styles import Font

    uid = request.args.get('user_id', type=int)
    mode = request.args.get('mode', 'week')
    offset = request.args.get('offset', 0, type=int)
    tz_hours = current_app.config.get('TZ_OFFSET_HOURS', 3)

    if not uid:
        uid = current_user.id
    if not current_user.can_admin and uid != current_user.id:
        return jsonify({'error': 'Нет прав'}), 403

    user = User.query.get_or_404(uid)
    start, end, period_label = _period_bounds(mode, offset, tz_hours)

    worked_task_ids = db.session.query(TimeLog.task_id).filter(
        TimeLog.user_id == uid,
        TimeLog.ended_at.isnot(None),
    ).distinct().subquery()
    done_tasks = Task.query.filter(
        Task.id.in_(worked_task_ids),
        Task.status == TaskStatus.DONE,
        Task.completed_at >= start,
        Task.completed_at <= end,
        Task.completed_at.isnot(None),
    ).order_by(Task.completed_at.desc()).all()

    logs = TimeLog.query.filter(
        TimeLog.user_id == uid,
        TimeLog.started_at >= start,
        TimeLog.started_at <= end,
        TimeLog.ended_at.isnot(None),
    ).all()
    time_by_task = defaultdict(float)
    for log in logs:
        time_by_task[log.task_id] += (log.ended_at - log.started_at).total_seconds()

    by_type = defaultdict(lambda: {'cnt': 0, 'secs': 0})
    for t in done_tasks:
        secs = int(time_by_task.get(t.id, 0))
        type_label = TYPE_LABELS.get(t.task_type, t.task_type or 'Не указан')
        by_type[type_label]['cnt'] += 1
        by_type[type_label]['secs'] += secs

    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = 'Задачи'
    ws1.append(['ID', 'Заголовок', 'Тип', 'Отдел', 'Закрыта', 'Время (мин)'])
    for cell in ws1[1]:
        cell.font = Font(bold=True)
    for t in done_tasks:
        secs = int(time_by_task.get(t.id, 0))
        ws1.append([
            t.id,
            t.title,
            TYPE_LABELS.get(t.task_type, t.task_type or ''),
            t.department.name if t.department else '',
            (t.completed_at + timedelta(hours=tz_hours)).strftime('%d.%m.%Y %H:%M') if t.completed_at else '',
            round(secs / 60),
        ])

    ws2 = wb.create_sheet('По типам')
    ws2.append(['Тип задачи', 'Кол-во задач', 'Время (мин)', 'Время (ч)'])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for type_label, data in sorted(by_type.items(), key=lambda x: x[1]['cnt'], reverse=True):
        ws2.append([
            type_label,
            data['cnt'],
            round(data['secs'] / 60),
            round(data['secs'] / 3600, 1),
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    safe_name = user.full_name.replace(' ', '_')
    mode_labels = {'day': 'день', 'week': 'неделя', 'month': 'месяц'}
    filename = f'stats_{safe_name}_{mode_labels.get(mode, mode)}.xlsx'
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


def _register_cyrillic_font():
    """Register DejaVuSans TTF for Cyrillic support in ReportLab. Returns font name to use."""
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os
    font_name = 'DejaVuSans'
    # Check if already registered
    try:
        pdfmetrics.getFont(font_name)
        return font_name
    except KeyError:
        pass
    # Search paths: system (Debian/Ubuntu after fonts-dejavu-core) + bundled fallback
    search_paths = [
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
        '/usr/share/fonts/dejavu/DejaVuSans.ttf',
        '/usr/share/fonts/TTF/DejaVuSans.ttf',
        os.path.join(current_app.root_path, 'static', 'fonts', 'DejaVuSans.ttf'),
    ]
    for path in search_paths:
        if os.path.exists(path):
            pdfmetrics.registerFont(TTFont(font_name, path))
            return font_name
    return 'Helvetica'  # fallback (no Cyrillic, but won't crash)


@analytics_bp.route('/analytics/export/pdf')
@login_required
def export_pdf():
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from sqlalchemy.orm import joinedload

    font_name = _register_cyrillic_font()

    # Суммарное время на задачу через SQL
    secs_sq = db.session.query(
        TimeLog.task_id,
        func.coalesce(func.sum(
            func.extract('epoch', TimeLog.ended_at - TimeLog.started_at)
        ), 0).label('secs')
    ).filter(TimeLog.ended_at.isnot(None)).group_by(TimeLog.task_id).subquery()

    rows = (db.session.query(Task, func.coalesce(secs_sq.c.secs, 0))
            .outerjoin(secs_sq, Task.id == secs_sq.c.task_id)
            .order_by(Task.created_at.desc())
            .all())

    cell_style = ParagraphStyle('cell', fontName=font_name, fontSize=7, leading=9)
    hdr_style  = ParagraphStyle('hdr',  fontName=font_name, fontSize=7, leading=9,
                                 textColor=colors.white)

    def p(text, style=None):
        return Paragraph(str(text), style or cell_style)

    headers = ['ID', 'Заголовок', 'Тип', 'Статус', 'Срочность', 'Заказчик', 'Дедлайн', 'Мин']
    data = [[p(h, hdr_style) for h in headers]]

    for t, total_secs in rows:
        data.append([
            p(t.id),
            p(t.title[:60]),
            p(TYPE_LABELS.get(t.task_type, t.task_type or '')[:25]),
            p(TaskStatus.LABELS.get(t.status, t.status)),
            p(Urgency.LABELS.get(t.urgency, t.urgency)),
            p(t.customer_name or ''),
            p(t.deadline.strftime('%d.%m.%Y') if t.deadline else ''),
            p(str(round(total_secs / 60))),
        ])

    col_widths = [30, 180, 120, 60, 60, 90, 60, 35]
    table = Table(data, repeatRows=1, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#343a40')),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4),
                            leftMargin=20, rightMargin=20,
                            topMargin=20, bottomMargin=20)
    doc.build([table])
    output.seek(0)
    return send_file(output, mimetype='application/pdf',
                     as_attachment=True, download_name='tasks_export.pdf')
