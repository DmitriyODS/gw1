import calendar
import io
from datetime import date, datetime
from flask import Blueprint, render_template, request, send_file
from flask_login import login_required
from sqlalchemy import or_, and_
from models import Task, TaskStatus

media_plan_bp = Blueprint('media_plan', __name__)

PUB_SUBTYPE_LABELS = {'news': 'Новость', 'event': 'Мероприятие'}
PLATFORM_ICONS = {
    'Сайт': 'bi-globe',
    'Внешние соц. сети': 'bi-share',
    'Внутренние соц. сети': 'bi-people',
    'Афиша': 'bi-megaphone',
}

MONTH_NAMES_RU = [
    '', 'Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
    'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь'
]


def _parse_pub_date(task):
    """Return datetime from dynamic_fields.pub_date or None."""
    raw = (task.dynamic_fields or {}).get('pub_date', '')
    if not raw:
        return None
    try:
        return datetime.fromisoformat(raw)
    except Exception:
        return None


def _get_month_tasks(current_month):
    """Return (by_day, tasks_no_date) for the given month.
    by_day: {day_int: [task, ...]} sorted by time within day.
    tasks_no_date: tasks without pub_date.
    """
    month_start = date(current_month.year, current_month.month, 1)
    if current_month.month == 12:
        month_end = date(current_month.year + 1, 1, 1)
    else:
        month_end = date(current_month.year, current_month.month + 1, 1)

    # Фильтруем в БД: не архивные, только нужный месяц или без даты
    # dynamic_fields['pub_date'].astext использует PostgreSQL оператор ->>
    pub_date_col = Task.dynamic_fields['pub_date'].astext
    all_pubs = Task.query.filter(
        Task.task_type.in_(['placement', 'publication']),
        Task.is_archived == False,
        or_(
            pub_date_col.is_(None),                          # ключ отсутствует в JSON
            pub_date_col == '',                              # пустая строка
            and_(
                pub_date_col >= month_start.isoformat(),
                pub_date_col < month_end.isoformat(),
            )
        )
    ).all()

    by_day = {}
    tasks_no_date = []

    for t in all_pubs:
        pub_dt = _parse_pub_date(t)
        if pub_dt is None:
            tasks_no_date.append(t)
        elif month_start <= pub_dt.date() < month_end:
            d = pub_dt.day
            # Attach parsed dt for sorting / display
            t.pub_dt = pub_dt
            by_day.setdefault(d, []).append(t)

    # Sort each day's tasks by time
    for d in by_day:
        by_day[d].sort(key=lambda t: t.pub_dt)

    return by_day, tasks_no_date


def _current_month_from_request():
    month_str = request.args.get('month')
    try:
        year, mon = (int(x) for x in month_str.split('-')) if month_str else (None, None)
        return date(year, mon, 1)
    except Exception:
        today = date.today()
        return date(today.year, today.month, 1)


@media_plan_bp.route('/media-plan')
@login_required
def index():
    current_month = _current_month_from_request()

    if current_month.month == 1:
        prev_month = f'{current_month.year - 1}-12'
    else:
        prev_month = f'{current_month.year}-{current_month.month - 1:02d}'
    if current_month.month == 12:
        next_month = f'{current_month.year + 1}-01'
    else:
        next_month = f'{current_month.year}-{current_month.month + 1:02d}'

    cal = calendar.monthcalendar(current_month.year, current_month.month)
    by_day, tasks_no_date = _get_month_tasks(current_month)

    return render_template('media_plan/index.html',
        current_month=current_month,
        month_label=f'{MONTH_NAMES_RU[current_month.month]} {current_month.year}',
        prev_month=prev_month,
        next_month=next_month,
        cal=cal,
        by_day=by_day,
        tasks_no_date=tasks_no_date,
        today=date.today(),
        TaskStatus=TaskStatus,
        PUB_SUBTYPE_LABELS=PUB_SUBTYPE_LABELS,
        PLATFORM_ICONS=PLATFORM_ICONS,
    )


@media_plan_bp.route('/media-plan/export')
@login_required
def export_xlsx():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    current_month = _current_month_from_request()
    by_day, tasks_no_date = _get_month_tasks(current_month)

    # Flatten sorted list
    all_tasks = []
    for day in sorted(by_day.keys()):
        for t in by_day[day]:
            all_tasks.append((t.pub_dt, t))
    for t in tasks_no_date:
        all_tasks.append((None, t))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{MONTH_NAMES_RU[current_month.month]} {current_month.year}'

    STATUS_LABELS = {
        'new': 'Новая', 'in_progress': 'В работе',
        'paused': 'На паузе', 'review': 'Проверка', 'done': 'Готово',
    }

    headers = ['Дата публикации', 'Время', 'Название', 'Подтип', 'Площадки', 'Статус', 'Ссылка']
    header_fill = PatternFill('solid', fgColor='6366F1')
    thin = Side(style='thin', color='DDDDDD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    ws.row_dimensions[1].height = 20

    STATUS_COLORS = {
        'new': 'E2E8F0', 'in_progress': 'DBEAFE',
        'paused': 'FEF9C3', 'review': 'DBEAFE', 'done': 'DCFCE7',
    }

    for row_idx, (pub_dt, t) in enumerate(all_tasks, 2):
        df = t.dynamic_fields or {}
        platforms = ', '.join(df.get('platforms') or [])
        subtype = PUB_SUBTYPE_LABELS.get(df.get('subtype', ''), df.get('subtype', '') or '')
        date_str = pub_dt.strftime('%d.%m.%Y') if pub_dt else '—'
        time_str = pub_dt.strftime('%H:%M') if pub_dt else '—'
        status_label = STATUS_LABELS.get(t.status, t.status)
        pub_url = df.get('pub_url', '')

        row_fill = PatternFill('solid', fgColor=STATUS_COLORS.get(t.status, 'FFFFFF'))
        values = [date_str, time_str, t.title, subtype, platforms, status_label, pub_url]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=(col == 3))
        ws.row_dimensions[row_idx].height = 18

    # Column widths
    for col, width in enumerate([14, 8, 40, 14, 30, 12, 40], 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"mediaplan_{current_month.year}_{current_month.month:02d}.xlsx"
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)
